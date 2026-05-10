# Prothomalo Scrapping Script

import os
import time
import re
from urllib.parse import urljoin
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import (
    NoSuchElementException,
    StaleElementReferenceException,
    ElementClickInterceptedException,
    WebDriverException
)
from bs4 import BeautifulSoup
import requests
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
from datetime import datetime

# =========================
# CONFIG
# =========================
TEST_MODE = False
TEST_LINK = "https://www.prothomalo.com/bangladesh/crime/iqnknzxbyq"
BASE_URL = "https://www.prothomalo.com/collection/latest"
SAVE_DIR = r"D:\No_2025\DR\Partho_Vai_September_13_2025\Scrapping_Test\Scrapping_Easy\Prothomalo\OandC\May_07_2026"

_ts = time.strftime("%Y%m%d_%H%M%S")
EXCEL_FILE = os.path.join(SAVE_DIR, f"ProthomAlo_{_ts}.xlsx")
SESSION_LOG_FILE = os.path.join(SAVE_DIR, "ProthomAlo_Session_Log.xlsx")

HEADLESS = False
MAX_LOAD_MORE = None
NO_NEW_AFTER_CLICKS = 4
MAX_SCROLL_ATTEMPTS = 20

ARTICLE_GROWTH_WAIT_SECONDS = 12
INITIAL_PAGE_LOAD_SLEEP = 2
POST_CLICK_SLEEP = 0.35
LAZY_SCROLL_SLEEP_1 = 1.0
LAZY_SCROLL_SLEEP_2 = 0.6

BATCH_EVERY_CLICKS = 40

# =========================
# SESSION LOGGING (START)
# one run = one session
# =========================
session_id = f"PROTHOMALO_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
source_name = "Prothom Alo"
start_time = datetime.now()

session_info = {
    "total_saved": 0,
    "total_processed": 0,
    "status": "running",
    "stop_reason": ""
}

# =========================
# TODAY-ONLY DATE FILTER
# =========================
TODAY_ISO = datetime.now().strftime("%Y-%m-%d")

def _iso_to_date(s):
    try:
        return datetime.strptime(str(s).strip(), "%Y-%m-%d").date()
    except Exception:
        return None

def _is_today(iso_date):
    d = _iso_to_date(iso_date)
    t = _iso_to_date(TODAY_ISO)
    return d is not None and t is not None and d == t

def _older_than_today(iso_date):
    d = _iso_to_date(iso_date)
    t = _iso_to_date(TODAY_ISO)
    return d is not None and t is not None and d < t

os.makedirs(SAVE_DIR, exist_ok=True)

# =========================
# INIT ARTICLE EXCEL
# =========================
COLUMNS = [
    "Newspaper", "Category", "Sub-Head", "Headline", "Sub-Head-2",
    "Link", "Author", "Location",
    "Published_Time", "Published_Time_Eng", "Published_Date",
    "Updated_Time", "Updated_Time_Eng",
    "Content", "Version"
]

if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(COLUMNS)
    wb.save(EXCEL_FILE)

# =========================
# INIT SESSION LOG EXCEL
# one row per run
# =========================
SESSION_LOG_COLUMNS = [
    "session_id", "source_name", "start_time", "end_time",
    "duration_seconds", "total_saved", "total_processed",
    "status", "stop_reason", "output_file"
]

if not os.path.exists(SESSION_LOG_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(SESSION_LOG_COLUMNS)
    wb.save(SESSION_LOG_FILE)

# =========================
# INIT SELENIUM
# =========================
options = Options()
if HEADLESS:
    options.add_argument("--headless=new")

options.add_argument("--start-maximized")
options.add_argument("--log-level=3")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 15)

# =========================
# HELPER: Author + Location extractor
# =========================
def extract_author_and_location(soup):
    author = ""
    location = ""

    wrapper = soup.find("div", class_="author-name-location-wrapper")
    if wrapper:
        name_span = wrapper.select_one("span.contributor-name")
        loc_span = wrapper.select_one("span.author-location")

        if name_span:
            data_author = name_span.get("data-author-0", "").strip()
            visible_text = name_span.text.strip()
            if visible_text == "লেখা:" and data_author:
                return data_author, ""
            author = data_author or visible_text

        if loc_span:
            location = loc_span.text.strip()

        if author:
            return author, location

    for sp in soup.select("span.contributor-name"):
        data_author = sp.get("data-author-0")
        if data_author and sp.text.strip() == "লেখা:":
            return data_author.strip(), ""

    sp = soup.select_one("span.contributor-name[data-author-0]")
    if sp:
        return sp.get("data-author-0").strip(), ""

    sp = soup.select_one("span.contributor-name")
    if sp:
        return sp.text.strip(), ""

    return "", ""

# =========================
# HELPER: Published + Updated Time
# =========================
def extract_published_and_updated_time(driver):
    def read_current_view():
        published_time = ""
        updated_time = ""
        for t in driver.find_elements(By.CSS_SELECTOR, "time"):
            txt = t.text.strip()
            if not txt:
                continue
            if "প্রকাশ" in txt:
                published_time = txt.replace("প্রকাশ:", "").strip()
            elif "আপডেট" in txt:
                updated_time = txt.replace("আপডেট:", "").strip()
        return published_time, updated_time

    published_time, updated_time = read_current_view()

    if updated_time and not published_time:
        try:
            toggle_div = driver.find_element(By.CSS_SELECTOR, "div.xuoYp")
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", toggle_div)
            driver.execute_script("arguments[0].click();", toggle_div)
            time.sleep(0.35)
            p2, u2 = read_current_view()
            if p2:
                published_time = p2
            if u2:
                updated_time = u2
        except Exception:
            pass

    return published_time, updated_time

# =========================
# HELPER: Bangla date -> English / ISO
# =========================
def extract_date_only(bn_datetime_text, output="en"):
    """
    output="en"  -> "16 February 2026"
    output="iso" -> "2026-02-16"
    """
    if not bn_datetime_text:
        return ""

    date_part = str(bn_datetime_text).split(",", 1)[0].strip()

    bn_to_en_digits = str.maketrans("০১২৩৪৫৬৭৮৯", "0123456789")
    date_part = date_part.translate(bn_to_en_digits)
    date_part = re.sub(r"\s+", " ", date_part).strip()

    month_map = {
        "জানুয়ারি": ("January", 1), "জানুয়ারি": ("January", 1),
        "ফেব্রুয়ারি": ("February", 2), "ফেব্রুয়ারি": ("February", 2),
        "মার্চ": ("March", 3),
        "এপ্রিল": ("April", 4),
        "মে": ("May", 5),
        "জুন": ("June", 6),
        "জুলাই": ("July", 7),
        "আগস্ট": ("August", 8),
        "সেপ্টেম্বর": ("September", 9),
        "অক্টোবর": ("October", 10),
        "নভেম্বর": ("November", 11),
        "ডিসেম্বর": ("December", 12),
    }

    m = re.search(r"(\d{1,2})\s+([^\s]+)\s+(\d{4})", date_part)
    if not m:
        return date_part

    day = int(m.group(1))
    month_bn = m.group(2).strip()
    year = int(m.group(3))

    if month_bn in month_map:
        month_en, month_num = month_map[month_bn]
        if output == "iso":
            return f"{year:04d}-{month_num:02d}-{day:02d}"
        return f"{day} {month_en} {year}"

    return date_part

# =========================
# HELPER: Convert Bangla datetime to English datetime
# Example:
# "১৩ এপ্রিল ২০২৬, ১১: ১৯" -> "13 April 2026, 11:19"
# =========================
def convert_bn_datetime_to_english(bn_datetime_text):
    if not bn_datetime_text:
        return ""

    try:
        text = str(bn_datetime_text).strip()

        bn_to_en_digits = str.maketrans("০১২৩৪৫৬৭৮৯", "0123456789")
        text = text.translate(bn_to_en_digits)

        text = re.sub(r"\s+", " ", text).strip()
        text = re.sub(r"\s*:\s*", ":", text)

        month_map = {
            "জানুয়ারি": "January", "জানুয়ারি": "January",
            "ফেব্রুয়ারি": "February", "ফেব্রুয়ারি": "February",
            "মার্চ": "March",
            "এপ্রিল": "April",
            "মে": "May",
            "জুন": "June",
            "জুলাই": "July",
            "আগস্ট": "August",
            "সেপ্টেম্বর": "September",
            "অক্টোবর": "October",
            "নভেম্বর": "November",
            "ডিসেম্বর": "December",
        }

        for bn_month, en_month in month_map.items():
            text = text.replace(bn_month, en_month)

        m = re.match(
            r"^\s*(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})(?:,\s*(\d{1,2}):(\d{1,2}))?\s*$",
            text
        )
        if m:
            day = int(m.group(1))
            month_en = m.group(2)
            year = int(m.group(3))
            hh = m.group(4)
            mm = m.group(5)

            if hh is not None and mm is not None:
                return f"{day} {month_en} {year}, {int(hh):02d}:{int(mm):02d}"
            return f"{day} {month_en} {year}"

        return text

    except Exception:
        return ""

# =========================
# SAFE ARTICLE EXCEL APPEND
# =========================
def append_to_excel(row, max_retries=3):
    for attempt in range(1, max_retries + 1):
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            ws.append(row)
            wb.save(EXCEL_FILE)
            return True
        except Exception as e:
            if attempt == max_retries:
                print(f"❌ Excel write failed after {max_retries} retries: {e}")
                return False
            time.sleep(0.6)

# =========================
# SAFE SESSION LOG APPEND
# one summary row per session
# =========================
def append_session_log(row, max_retries=3):
    for attempt in range(1, max_retries + 1):
        try:
            wb = load_workbook(SESSION_LOG_FILE)
            ws = wb.active
            ws.append(row)
            wb.save(SESSION_LOG_FILE)
            return True
        except Exception as e:
            if attempt == max_retries:
                print(f"❌ Session log write failed after {max_retries} retries: {e}")
                return False
            time.sleep(0.6)

# =========================
# SAVE SESSION SUMMARY
# called once at end
# =========================
def save_session_summary(end_time):
    duration_seconds = int((end_time - start_time).total_seconds())

    row = [
        session_id,
        source_name,
        start_time.strftime("%Y-%m-%d %H:%M:%S"),
        end_time.strftime("%Y-%m-%d %H:%M:%S"),
        duration_seconds,
        session_info["total_saved"],
        session_info["total_processed"],
        session_info["status"],
        session_info["stop_reason"],
        EXCEL_FILE
    ]
    append_session_log(row)

# =========================
# PRINT SESSION SUMMARY
# notebook-friendly output
# =========================
def print_session_summary(end_time):
    duration_seconds = int((end_time - start_time).total_seconds())

    print("\n" + "=" * 60)
    print("SCRAPING SESSION SUMMARY")
    print("=" * 60)
    print(f"Session ID      : {session_id}")
    print(f"Source Name     : {source_name}")
    print(f"Start Time      : {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"End Time        : {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Duration (sec)  : {duration_seconds}")
    print(f"Total Saved     : {session_info['total_saved']}")
    print(f"Total Processed : {session_info['total_processed']}")
    print(f"Status          : {session_info['status']}")
    print(f"Stop Reason     : {session_info['stop_reason']}")
    print(f"Output File     : {EXCEL_FILE}")
    print(f"Session Log File: {SESSION_LOG_FILE}")
    print("=" * 60)

# =========================
# SAFE REQUESTS SESSION
# =========================
def make_session():
    session = requests.Session()
    retries = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"]
    )
    adapter = HTTPAdapter(max_retries=retries)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome Safari/537.36"
        )
    })
    return session

# =========================
# FETCH ARTICLE DETAILS
# =========================
def fetch_article_details(link, session):
    newspaper = "Prothom Alo"
    category = sub_head = headline = sub_head_2 = author = location = ""
    published_time = published_time_eng = updated_time = updated_time_eng = content_text = version = ""
    published_date = ""

    try:
        driver.execute_script("window.open('');")
        driver.switch_to.window(driver.window_handles[-1])
        driver.get(link)
        time.sleep(0.8)

        try:
            headline_elem = driver.find_element(By.CSS_SELECTOR, "h1")
            headline = headline_elem.text.strip()
            if not headline:
                raise Exception("No headline")
        except Exception:
            try:
                driver.close()
            except Exception:
                pass
            driver.switch_to.window(driver.window_handles[0])
            return None

        try:
            sub_head_elem = driver.find_element(By.CSS_SELECTOR, "div.print-entity-section-wrapper h2")
            sub_head = sub_head_elem.text.strip()
        except NoSuchElementException:
            sub_head = ""

        try:
            cat_elem = driver.find_element(By.CSS_SELECTOR, "a[itemprop='genre']")
            category = cat_elem.text.strip()
        except Exception:
            match = re.search(r"prothomalo\.com/([^/]+)/", link)
            category = match.group(1) if match else ""

        try:
            published_time, updated_time = extract_published_and_updated_time(driver)
            published_time_eng = convert_bn_datetime_to_english(published_time)
            published_date = extract_date_only(published_time, output="iso")
            updated_time_eng = convert_bn_datetime_to_english(updated_time)
        except Exception:
            published_time = published_time_eng = updated_time = updated_time_eng = ""
            published_date = ""

        try:
            driver.close()
        except Exception:
            pass
        driver.switch_to.window(driver.window_handles[0])

    except Exception as e:
        try:
            driver.close()
        except Exception:
            pass
        try:
            driver.switch_to.window(driver.window_handles[0])
        except Exception:
            pass
        print(f"Selenium error: {e}")

    try:
        resp = session.get(link, timeout=12)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.content, "html.parser")

        article = (
            soup.find("div", class_="story-content")
            or soup.find("div", class_="article-body")
            or soup.find("article")
        )
        if article:
            paragraphs = article.find_all("p")
            content_text = "\n".join(
                p.get_text(strip=True) for p in paragraphs if p.get_text(strip=True)
            )
        else:
            content_text = ""

        author, location = extract_author_and_location(soup)
        version = ""

    except Exception as e:
        print(f"Requests/BS error: {e}")

    return [
        newspaper, category, sub_head, headline, sub_head_2,
        link, author, location,
        published_time, published_time_eng, published_date,
        updated_time, updated_time_eng,
        content_text, version
    ]

# =========================
# Collect current ordered unique links from listing
# =========================
def get_current_listing_links():
    article_elements = driver.find_elements(By.CSS_SELECTOR, "h3.headline-title a")
    ordered = []
    seen = set()

    for a in article_elements:
        href = a.get_attribute("href")
        if href:
            full = urljoin("https://www.prothomalo.com", href)
            if full not in seen:
                seen.add(full)
                ordered.append(full)

    return ordered

# =========================
# Console print for article row
# =========================
def print_article_console(article_no, row):
    content = row[13] or ""
    if len(content) > 80:
        content_preview = content[:40] + " ... " + content[-40:]
    else:
        content_preview = content

    print(f"\nArticle {article_no}")
    print(f"Newspaper: {row[0]}")
    print(f"Category: {row[1]}")
    print(f"Sub-Head: {row[2]}")
    print(f"Headline: {row[3]}")
    print(f"Link: {row[5]}")
    print(f"Author: {row[6]}")
    print(f"Location: {row[7]}")
    print(f"Published_Time: {row[8]}")
    print(f"Published_Time_Eng: {row[9]}")
    print(f"Published_Date: {row[10]}")
    print(f"Updated_Time: {row[11]}")
    print(f"Updated_Time_Eng: {row[12]}")
    print(f"Content (preview): {content_preview}")
    print(f"Version: {row[14]}")

# =========================
# Scrape newly discovered links and save immediately
# =========================
def scrape_new_links(new_links, session, processed_links, saved_counter_ref, stop_ref):
    for link in new_links:
        if not link or link in processed_links:
            continue

        # count every link processed in this run
        session_info["total_processed"] += 1

        try:
            row = fetch_article_details(link, session)
            if row:
                pub_iso = (row[10] or "").strip()

                # no valid date => skip safely
                if not _iso_to_date(pub_iso):
                    processed_links.add(link)
                    continue

                # older than today => stop whole process safely
                if _older_than_today(pub_iso):
                    print(f"🛑 Reached older news. pub={pub_iso} < today={TODAY_ISO}. Stopping safely.")
                    session_info["stop_reason"] = "Reached older news"
                    stop_ref["stop"] = True
                    return

                # save only today's articles
                if not _is_today(pub_iso):
                    processed_links.add(link)
                    continue

                ok = append_to_excel(row)
                if ok:
                    processed_links.add(link)
                    saved_counter_ref["n"] += 1
                    session_info["total_saved"] = saved_counter_ref["n"]
                    print_article_console(saved_counter_ref["n"], row)

        except Exception as e:
            print(f"Error processing article {link}: {e}")
            continue

# =========================
# MAIN
# =========================
try:
    session = make_session()
    processed_links = set()
    seen_listing_links = set()
    saved_counter = {"n": 0}
    stop_flag = {"stop": False}

    if TEST_MODE:
        session_info["stop_reason"] = "Test mode"
        print("🔬 TEST_MODE: scraping only 1 article")
        scrape_new_links([TEST_LINK], session, processed_links, saved_counter, stop_flag)

    else:
        driver.get(BASE_URL)
        time.sleep(INITIAL_PAGE_LOAD_SLEEP)

        print(f"🎯 Today-only scraping active: {TODAY_ISO}")
        print("Starting lazy-scroll to load visible articles (before clicking 'আরও') ...")

        last_count = 0
        scroll_attempt = 0

        while scroll_attempt < MAX_SCROLL_ATTEMPTS:
            try:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            except WebDriverException:
                pass

            time.sleep(LAZY_SCROLL_SLEEP_1)
            time.sleep(LAZY_SCROLL_SLEEP_2)

            article_elements = driver.find_elements(By.CSS_SELECTOR, "h3.headline-title a")
            current_count = len(article_elements)
            print(f"  lazy-scroll attempt {scroll_attempt + 1}: found {current_count} article links so far")

            if current_count == last_count:
                break

            last_count = current_count
            scroll_attempt += 1

        current_links = get_current_listing_links()
        new_links = [l for l in current_links if l not in seen_listing_links]
        seen_listing_links.update(new_links)

        if new_links:
            print(f"✅ Initial batch: scraping {len(new_links)} newly loaded links...")
            scrape_new_links(new_links, session, processed_links, saved_counter, stop_flag)
            if stop_flag["stop"]:
                raise SystemExit

        print("Attempting to click 'আরও' and scrape in batches (safe) ...")
        load_more_attempt = 0
        no_new_clicks = 0

        while True:
            if stop_flag["stop"]:
                break

            if MAX_LOAD_MORE is not None and load_more_attempt >= MAX_LOAD_MORE:
                session_info["stop_reason"] = f"Reached MAX_LOAD_MORE = {MAX_LOAD_MORE}"
                print(f"  Reached MAX_LOAD_MORE = {MAX_LOAD_MORE}. Stopping.")
                break

            try:
                load_more = driver.find_element(By.CSS_SELECTOR, "span.load-more-content")
            except NoSuchElementException:
                session_info["stop_reason"] = "'আরও' button not found"
                print("  'আরও' button not found (no more load-more).")
                break

            prev_count = len(driver.find_elements(By.CSS_SELECTOR, "h3.headline-title a"))

            try:
                driver.execute_script(
                    "arguments[0].scrollIntoView({behavior:'smooth', block:'center'});",
                    load_more
                )
            except WebDriverException:
                pass

            time.sleep(POST_CLICK_SLEEP)

            clicked = False
            try:
                load_more.click()
                clicked = True
                print(f"  Clicked 'আরও' (attempt {load_more_attempt + 1}) via element.click()")
            except (ElementClickInterceptedException, StaleElementReferenceException, WebDriverException):
                try:
                    driver.execute_script("arguments[0].click();", load_more)
                    clicked = True
                    print(f"  Clicked 'আরও' (attempt {load_more_attempt + 1}) via JS click fallback")
                except Exception as e:
                    session_info["stop_reason"] = f"Failed to click 'আরও': {e}"
                    print(f"  Failed to click 'আরও' on attempt {load_more_attempt + 1}: {e}")
                    break

            if not clicked:
                session_info["stop_reason"] = "Load more click failed"
                break

            waited = 0
            grew = False
            while waited < ARTICLE_GROWTH_WAIT_SECONDS:
                time.sleep(1)
                curr_count = len(driver.find_elements(By.CSS_SELECTOR, "h3.headline-title a"))
                if curr_count > prev_count:
                    print(f"    new articles loaded: {curr_count} total")
                    grew = True
                    break
                waited += 1

            if not grew:
                no_new_clicks += 1
                print(f"    No additional articles loaded (no_new_clicks={no_new_clicks}).")
                if no_new_clicks >= NO_NEW_AFTER_CLICKS:
                    session_info["stop_reason"] = "'আরও' stopped producing new articles"
                    print("    Stopping safely: 'আরও' stopped producing new articles.")
                    break
            else:
                no_new_clicks = 0

            load_more_attempt += 1

            if load_more_attempt % BATCH_EVERY_CLICKS == 0:
                current_links = get_current_listing_links()
                new_links = [l for l in current_links if l not in seen_listing_links]
                seen_listing_links.update(new_links)

                if new_links:
                    print(f"✅ Batch after {load_more_attempt} clicks: scraping {len(new_links)} new links...")
                    scrape_new_links(new_links, session, processed_links, saved_counter, stop_flag)
                    if stop_flag["stop"]:
                        break
                else:
                    print(f"ℹ️ Batch after {load_more_attempt} clicks: no new links found.")

        if not stop_flag["stop"]:
            final_links = get_current_listing_links()
            final_new_links = [l for l in final_links if l not in seen_listing_links]
            seen_listing_links.update(final_new_links)

            if final_new_links:
                print(f"✅ Final batch: scraping {len(final_new_links)} remaining new links...")
                scrape_new_links(final_new_links, session, processed_links, saved_counter, stop_flag)

        if not session_info["stop_reason"]:
            session_info["stop_reason"] = "Completed normally"

        print(f"\n✅ Done. Total today's articles saved: {saved_counter['n']}")
        print(f"✅ File saved at: {EXCEL_FILE}")

    # run completed with some or no saved data
    session_info["status"] = "success" if session_info["total_saved"] > 0 else "partial_success"

except Exception as e:
    session_info["status"] = "failed"
    if not session_info["stop_reason"]:
        session_info["stop_reason"] = f"Unhandled error: {e}"
    print(f"\n❌ Fatal error: {e}")

finally:
    end_time = datetime.now()

    try:
        print_session_summary(end_time)
    except Exception as e:
        print(f"❌ Failed to print session summary: {e}")

    try:
        save_session_summary(end_time)
        print(f"✅ Session log saved at: {SESSION_LOG_FILE}")
    except Exception as e:
        print(f"❌ Failed to save session log: {e}")

    try:
        driver.quit()
    except Exception:
        pass

